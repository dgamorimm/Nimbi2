from selenium import webdriver
from modelo.login import User
from modelo.requisicao import Requisicao
from processador.log import Logger
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
import os
import shutil
from time import sleep

logger = Logger().get()
logger.info('log ON')

path_file = "c:\\Nimbi\\Downloads"
logger.info("Variavel = path_file ------> Passando caminho para download")

prefs = {"safebrowsing.enabled": True}
prefs2 = {"download.default_directory": path_file}
logger.info("Variavel = prefs ------> Passando dicionario de preferência de download sem confirmação")
logger.info("Variavel = prefs2 ------> Passando dicionario de preferencia de download no caminho da variavel path file")

prefs.update(prefs2)
logger.info("Funcao = prefs.update(prefs2) ------> Atualizando a variavel prefs, passando o dicionario prefs2")

options = webdriver.ChromeOptions()
logger.info("Variavel = options ------> Passando o webdriver das opções de driver do Google Chrome")

options.add_argument("-incognito")
logger.info("Funcao = options.add_argument('-incognito') ------> Adicionando o argumento a options do drive do Google")

options.add_argument("--disable-dev-shm-usage")
logger.info("Funcao = options.add_argument('--disable-dev-shm-usage') ------> Adicionando o argumento a options do drive do Google")

options.add_argument("start-maximized")
logger.info("Funcao = options.add_argument('start-maximized')------> Adicionando o argumento a options do drive do Google")

options.add_argument("--no-sandbox")
logger.info("Funcao = options.add_argument('--no-sandbox')------> Adicionando o argumento a options do drive do Google")
#options.add_argument("headless") - Ambos servem para rodar em background
#options.add_argument("disable-gpu")
#options.binary_location = "C:\\Users\\douglas.pinheiro\\AppData\\Local\\Google\\Chrome SxS\\Application\\chrome.exe"
options.add_experimental_option("prefs", prefs)
logger.info("Funcao = options.add_experimental_option('prefs', prefs)------> Adicionando o argumento a options do drive do Google")

bot = webdriver.Chrome(options=options, port=5556)
logger.info("Variavel = bot ------> Instanciando o driver do Google")

bot.get("https://nimbi.com.br/portais/estacio/")
logger.info("Funcao = bot.get ------> Abrindo a pagina da NIMBI")

usuario = User(bot)
logger.info("Classe = usuario ------> Instaciando a Classe User")

usuario.logar("EST1056883", "Ana@2018")
logger.info("Funcao = usuario.logar ------> Colocando usuário e senha nos argumentos")

req = Requisicao(bot)
logger.info("Classe = req  ------> Instaciando a Classe Requisicao")

caminho = '../dados/Teste.xlsx'
logger.info("Variavel = caminho  ------> Passando o caminho da planilha com as requisicoes")

arquivo_excel = load_workbook(caminho)
logger.info("Variavel = arquivo_excel  ------> Carregando a planilha in memory")

planilha1 = arquivo_excel.active
logger.info("Variavel = planilha1  ------> Ativando o sheet ( ' Planilha1 ')")

qtd_line = planilha1.max_row
logger.info("Variavel = qtd_line  ------> Contando a quantidade linhas com celulas preenchidas")

for i in range(2, qtd_line + 1):
    s_id_req = planilha1.cell(row=i, column=1).value
    logger.info("Variavel = s_id_req  ------> Passando o valor da celula A2")

    s_status_req = planilha1.cell(row=i, column=3).value
    logger.info("Variavel = s_id_req  ------> Passando o valor da celula C2")

    logger.info("Funcao = req.buscar_requisicao(s_id_req, s_status_req) ------> Ativada")
    req.buscar_requisicao(s_id_req, s_status_req)
    logger.info("Funcao = req.buscar_requisicao(s_id_req, s_status_req) ------> Concluida")
    sleep(5)

    dir_origem = 'C:/Nimbi/Downloads/'
    logger.info("Variavel = dir_origem ------> Passando o caminho da origem dos downloads")

    dir_destino = 'C:/Nimbi/Requisição/' + str(s_id_req)
    logger.info("Variavel = dir_destino ------> Destino dos downloads")

    # anexo na guia de item
    logger.info("Variavel = anexs ------> Ativada")
    anexs = bot.find_elements_by_class_name("x-grid3-col-ATTACHMENT_CLIP")
    for anex in anexs:
        if anex.text != "":
            anex.click()
            anexo = bot.find_element_by_css_selector("td > div > div.x-panel-bwrap > div.x-panel-body.x-panel-body-"
                                                     "noheader.x-form-label-left > div:nth-child(2) > div.x-edit-grid.x-"
                                                     "grid-panel.x-component.x-border > div.x-grid3 > div.x-grid3-viewport"
                                                     " > div.x-grid3-scroller > div > div")
            if anexo.text != " ":
                try:
                    logger.info("Funcao = req.download_guia_itens() ------> Ativada")
                    req.download_guia_itens()
                    logger.info("Funcao = req.download_guia_itens() ------> Concluida")
                    sleep(1)
                    shutil.rmtree(dir_destino, ignore_errors=True)
                    os.mkdir(dir_destino)
                    busca_arquivo = os.listdir(dir_origem)
                    for aquivo in busca_arquivo:
                        if aquivo.endswith('.zip'):
                            shutil.move(os.path.join(dir_origem, aquivo), os.path.join(dir_destino, aquivo))
                        else:
                            print("Não existe anexo para a requisição {}".format(str(s_id_req)))
                except NoSuchElementException as x:
                    print(x)
                    break
            break
        break
    logger.info("Variavel = anexs ------> Concluida")
    try:
        bot.find_element_by_xpath("//div[@class='webb-req-search-fields-panel-label webb-cart-item-label webb-left x-component']")
    except NoSuchElementException as x:
        print(x)
        logger.info("Funcao =  req.download_guia_anexos() ------> Ativada")
        req.download_guia_anexos()
        logger.info("Funcao =  req.download_guia_anexos() ------> Concluida")
        if os.path.exists(dir_destino + "/anexos.zip"):
            os.rename("C:/Nimbi/Downloads/anexos.zip","C:/Nimbi/Downloads/anexos (1).zip")
            busca_arquivo = os.listdir(dir_origem)
            for aquivo in busca_arquivo:
                if aquivo.endswith('.zip'):
                    shutil.move(os.path.join(dir_origem, aquivo), os.path.join(dir_destino, aquivo))
        else:
            shutil.rmtree(dir_destino, ignore_errors=True)
            os.mkdir(dir_destino)
            if os.path.exists("C:/Nimbi/Downloads/anexos.zip"):
                busca_arquivo = os.listdir(dir_origem)
                for aquivo in busca_arquivo:
                    if aquivo.endswith('.zip'):
                        shutil.move(os.path.join(dir_origem, aquivo), os.path.join(dir_destino, aquivo))
            else:
                print("Não existe anexo para a requisição {}".format(str(s_id_req)))
    bot.refresh()
    sleep(2)



