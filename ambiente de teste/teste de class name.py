from selenium import webdriver
from modelo.login import User
from modelo.requisicao import Requisicao
from openpyxl import load_workbook
import os
import shutil
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

path_file = "c:\\Nimbi\\Requisição"


# Selecionar a aba de requisição
CLICK_REQ = (By.ID, "requisition-tab")
# Inclusão do ID
IN_REQ_ID = (By.NAME, "REQUISITION_WEBB_ID_NAME")  # "x-auto-1553-input")
# Selecionar o status
CLICK_STATUS = (By.NAME, "REQUISITION_STATUS_NAME")  # "x-auto-1555-input")
SELECT_EAPROVACAO = (By.XPATH, "//*[@id='x-auto-1556']/div[3]")
SELECT_APROVADO = (By.XPATH, "//*[@id='x-auto-1556']/div[4]")
SELECT_RECUSADO = (By.XPATH, "//*[@id='x-auto-1556']/div[5]")
SELECT_DEVOLVIDO = (By.XPATH, "//*[@id='x-auto-1556']/div[6]")
SELECT_RECEBIDO = (By.XPATH, "//*[@id='x-auto-1556']/div[7]")
SELECT_CANCELADO = (By.XPATH, "//*[@id='x-auto-1556']/div[8]")
# criado por todos
CLICK_CRIADOPOR = (By.ID, "x-auto-1562-input")
BTN_CRIADOPOR_TODOS = (By.TAG_NAME, "label[name='0'][title='Todos']")
# limpar criação de data
IN_DATAINIT = (By.ID, "start-date-input")
# buscar
ENTER_BUSCAR = (By.ID, "end-date-input")
# entrando pelo icone de folder
CLICK_ENTRAR = (By.NAME, "REQUISITION_OPEN_IMAGE_0_10_NAME")
# clicando no icone de anexo
CLICK_ANEXO = (By.TAG_NAME, "img[src='skin-webb/img/anexo_vazio.gif'][title='Anexos']")
# fazendo o download
ENTER_BAIXAR = (By.CSS_SELECTOR,"td:nth-child(1) > div > table > tbody > tr:nth-child(1) > td > div > table > tbody "
                     "> tr > td:nth-child(2) > table > tbody > tr:nth-child(2) > td.x-btn-mc > em >"
                     " button[type='button']")  # "x-window-body")#(By.ID,"x-auto-6061")

# fechar a tela de download
CLICK_INICIO = (By.ID, "home-tab")
# voltar para a pesquisa
CLICK_SETA_VOLTAR = (By.ID, "x-auto-779")

# Habilitando as configuracoes:
# Primeiro o diretorio que vai baixar o arquivo.
# Segundo remover a mensagem de permitir download.
prefs = {"safebrowsing.enabled": True}
prefs2 = {"download.default_directory": path_file}

prefs.update(prefs2)

options = webdriver.ChromeOptions()
options.add_argument("-incognito")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("start-maximized")
options.add_argument("--no-sandbox")
options.add_experimental_option("prefs", prefs)
print(options.experimental_options)
bot = webdriver.Chrome(options=options, port=5556)

bot.get("https://nimbi.com.br/portais/estacio/")

usuario = User(bot)

usuario.logar("EST1056883","Ana@2018")

req = Requisicao(bot)

caminho = '../dados/RelatorioRequisicao.xlsx'
arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active
qtd_line = planilha1.max_row

for i in range(2, qtd_line + 1):
    s_id_req = planilha1.cell(row=i, column=1).value
    s_status_req = planilha1.cell(row=i, column=3).value

    bot.switch_to.frame("topMenu")

    bot.find_element_by_id("requisition-tab").click()
    sleep(3)

    bot.switch_to.parent_frame()
    bot.switch_to.frame("gwtIndex")

    requis = bot.find_element_by_name("REQUISITION_WEBB_ID_NAME")
    requis.click()
    requis.clear()
    requis.send_keys(str(s_id_req))
    sleep(1)

    bot.find_element_by_name("REQUISITION_STATUS_NAME").click()
    sleep(1)

    status_list = bot.find_elements_by_class_name("x-combo-list-item ")

    for status in status_list:
        print(status.text)
    sleep(2)