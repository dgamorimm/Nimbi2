from openpyxl import load_workbook

caminho = './POLOS.xlsx'
arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active
qtd_line = planilha1.max_row

for i in range(2, qtd_line + 1):
    matriz = planilha1.cell(row=i, column=5).value
    matriz_subst = matriz.replace("EAD","EAD -").replace("/","-")
    matriz_list = matriz_subst.split("-")
    print(matriz_list[1].strip())
